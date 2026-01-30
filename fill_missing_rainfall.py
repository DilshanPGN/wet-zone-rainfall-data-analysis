"""
Fill missing rainfall data using the Normal Ratio Method (standard in hydrology).
Writes a new Excel file with imputed values and highlights imputed cells.

Method, equations, and step-by-step description: see docs/MISSING_DATA_IMPUTATION.md
"""
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# -----------------------------------------------------------------------------
# Configuration
# -----------------------------------------------------------------------------
DATASET_DIR = Path(__file__).resolve().parent / "dataset"
EXCEL_INPUT = DATASET_DIR / "Wet zone rainfall data.xlsx"
# Output file (avoids overwriting original if Excel has it open)
EXCEL_OUTPUT = DATASET_DIR / "Wet zone rainfall data_filled.xlsx"
SHEET_NAME = "All"
# Columns that may contain missing values to be filled (rainfall stations)
RAINFALL_COLUMNS = [
    "Colombo", "Galle", "Nuwara Eliya", "Rathnapura", "Maliboda", "Deniyaya"
]
# Highlight color for imputed cells (light yellow)
IMPUTED_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")


def compute_monthly_means(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """
    Compute long-term mean rainfall per month for each station.
    Returns DataFrame index = Month (1..12), columns = station names.
    """
    monthly_means = {}
    for col in columns:
        by_month = df.groupby("Month")[col].mean()
        monthly_means[col] = by_month
    return pd.DataFrame(monthly_means)


def normal_ratio_estimate(
    target_col: str,
    ref_value: float,
    ref_col: str,
    monthly_means: pd.DataFrame,
    month: int,
) -> float:
    """
    Single-reference Normal Ratio estimate:
        P_target = (mean_target_month / mean_ref_month) * P_ref
    """
    mean_target = monthly_means.loc[month, target_col]
    mean_ref = monthly_means.loc[month, ref_col]
    if mean_ref == 0:
        return 0.0
    return (mean_target / mean_ref) * ref_value


def fill_missing_with_normal_ratio(
    df: pd.DataFrame,
    target_col: str,
    monthly_means: pd.DataFrame,
    ref_columns: list[str],
) -> tuple[pd.DataFrame, pd.Series]:
    """
    Fill missing values in target_col using Normal Ratio with all reference stations.
    When multiple references have data, use the average of their estimates.
    Returns (df with filled values, boolean mask of rows that were imputed).
    """
    df = df.copy()
    imputed_mask = pd.Series(False, index=df.index)

    for idx in df.index:
        if pd.notna(df.loc[idx, target_col]):
            continue
        month = int(df.loc[idx, "Month"])
        estimates = []
        for ref_col in ref_columns:
            if ref_col == target_col:
                continue
            ref_val = df.loc[idx, ref_col]
            if pd.notna(ref_val) and ref_val >= 0:
                est = normal_ratio_estimate(
                    target_col, ref_val, ref_col, monthly_means, month
                )
                if pd.notna(est) and est >= 0:
                    estimates.append(est)
        if estimates:
            df.loc[idx, target_col] = sum(estimates) / len(estimates)
            imputed_mask.loc[idx] = True
        else:
            # Fallback: same-station linear interpolation (by index)
            # Use only this column's known values
            pass  # leave for second pass

    return df, imputed_mask


def fill_remaining_with_interpolation(
    df: pd.DataFrame,
    target_col: str,
    imputed_mask: pd.Series,
) -> tuple[pd.DataFrame, pd.Series]:
    """
    Fill any remaining NaN in target_col using linear interpolation along the series.
    """
    df = df.copy()
    still_missing = df[target_col].isna()
    if not still_missing.any():
        return df, imputed_mask
    # Interpolate in place; limit_direction='both' fills leading/trailing NaNs
    df.loc[:, target_col] = df[target_col].interpolate(
        method="linear", limit_direction="both"
    )
    newly_filled = still_missing & df[target_col].notna()
    imputed_mask = imputed_mask | newly_filled
    return df, imputed_mask


def run_imputation(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, pd.Series]]:
    """
    Run full imputation: monthly means, then Normal Ratio for each rainfall column,
    then interpolation fallback. Returns (filled DataFrame, {column: imputed mask}).
    """
    monthly_means = compute_monthly_means(df, RAINFALL_COLUMNS)
    ref_columns = RAINFALL_COLUMNS
    imputed_masks = {}

    for target_col in RAINFALL_COLUMNS:
        if not df[target_col].isna().any():
            imputed_masks[target_col] = pd.Series(False, index=df.index)
            continue
        df, imputed = fill_missing_with_normal_ratio(
            df, target_col, monthly_means, ref_columns
        )
        df, imputed = fill_remaining_with_interpolation(df, target_col, imputed)
        imputed_masks[target_col] = imputed

    return df, imputed_masks


def get_excel_cell_for_column(col_name: str, header_row: list) -> int:
    """Return 1-based column index (for openpyxl) for the given column name."""
    try:
        return header_row.index(col_name) + 1
    except ValueError:
        return -1


def apply_filled_data_to_excel(
    excel_input_path: Path,
    excel_output_path: Path,
    sheet_name: str,
    df_filled: pd.DataFrame,
    imputed_masks: dict[str, pd.Series],
) -> None:
    """
    Open the workbook, write imputed values only where imputed_masks is True,
    and set highlight fill on those cells. Saves to excel_output_path.
    """
    wb = load_workbook(excel_input_path)
    ws = wb[sheet_name]
    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    for col_name in RAINFALL_COLUMNS:
        col_idx = get_excel_cell_for_column(col_name, header_row)
        if col_idx < 0:
            continue
        mask = imputed_masks[col_name]
        for row_offset, was_imputed in enumerate(mask):
            if not was_imputed:
                continue
            # DataFrame row 0 = Excel row 2 (header in row 1)
            excel_row = row_offset + 2
            value = df_filled.loc[df_filled.index[row_offset], col_name]
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.value = round(float(value), 2)
            cell.fill = IMPUTED_FILL

    wb.save(excel_output_path)
    wb.close()


def main() -> None:
    if not EXCEL_INPUT.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_INPUT}")

    print("Loading data...")
    df = pd.read_excel(EXCEL_INPUT, sheet_name=SHEET_NAME)

    print("Running imputation (Normal Ratio + interpolation fallback)...")
    df_filled, imputed_masks = run_imputation(df)

    total_imputed = sum(imputed_masks[c].sum() for c in RAINFALL_COLUMNS)
    print(f"Total cells imputed: {total_imputed}")
    for col in RAINFALL_COLUMNS:
        n = imputed_masks[col].sum()
        if n > 0:
            print(f"  {col}: {n} cells")

    print("Writing Excel with highlighted imputed cells...")
    apply_filled_data_to_excel(
        EXCEL_INPUT, EXCEL_OUTPUT, SHEET_NAME, df_filled, imputed_masks
    )
    print(f"Done. Output file: {EXCEL_OUTPUT}")


if __name__ == "__main__":
    main()
